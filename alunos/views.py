import csv
import io
import os
import json
import openpyxl
from datetime import datetime
from django.utils import timezone
from urllib.parse import quote


from django.db.models import Q, Count, OuterRef
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse

from .models import Aluno, ArquivoAluno, WebSocialMember
from .forms import AlunoForm, AuxiliarAlunoForm, AlunoCSVUploadForm, VerificarCPFForm
from core.mixins import StaffRequiredMixin, AuditLogMixin
from escolas.models import Escola
from cursos.models import TipoCurso # Para cursos de interesse, se for o caso

def download_model_xlsx(request):
    # ... (content remains same, just context for replace)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modelo_importacao_alunos.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alunos"

    columns = [
        "escola_nome", "Nome", "Data nasc (00/00/0000)", "Cor", "Sexo", "Estado Civil", 
        "Filiacao", "Naturalidade", "uf", "cep", "endereco", "num", "bairro", "cidade", "estado_endereco",
        "telefone1", "email", "rg", "orgaoemissor", "dataemissao", "cpf", 
        "deficiencia (sim/nao)", "escolaridade", "situacao", "renda", 
        "nummoradores", "numtrabalham", "rendatotal", "valorresidencia",
        "tempo_moradia", "tipo_moradia", "como_soube"
    ]
    
    ws.append(columns)
    
    wb.save(response)
    return response


class SuperuserRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser

class AlunoVerificarCPFView(StaffRequiredMixin, View):
    template_name = 'alunos/verificar_cpf.html'
    
    def get(self, request):
        form = VerificarCPFForm()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        form = VerificarCPFForm(request.POST)
        if form.is_valid():
            cpf = form.cleaned_data['cpf']
            user = request.user
            
            if user.is_superuser:
                # Superuser creates for any school? Usually superuser flow is different, 
                # but let's assume they can proceed to create.
                # For simplicity, superuser just goes to create view with CPF
                return redirect(f"{reverse_lazy('alunos:criar_aluno')}?cpf={cpf}")

            if not hasattr(user, 'profile') or not user.profile.escola:
                messages.error(request, "Você não está vinculado a nenhuma escola.")
                return redirect('home')

            escola_atual = user.profile.escola

            # Check if exists in current school
            if Aluno.objects.filter(escola=escola_atual, cpf=cpf).exists():
                messages.error(request, f"O Aluno com CPF {cpf} já está cadastrado nesta escola ({escola_atual.nome}).")
                return render(request, self.template_name, {'form': form})
            
            # Check if exists in ANY other school
            aluno_existente = Aluno.objects.filter(cpf=cpf).first()
            if aluno_existente:
                # Found in another school -> Offer to clone
                return render(request, self.template_name, {
                    'form': form,
                    'aluno_existente': aluno_existente,
                    'mostrar_opcao_clonar': True
                })
            
            # Not found anywhere -> Redirect to Create with CPF pre-filled
            return redirect(f"{reverse_lazy('alunos:criar_aluno')}?cpf={cpf}")
            
        return render(request, self.template_name, {'form': form})

class AlunoClonarView(AuditLogMixin, StaffRequiredMixin, View):
    def post(self, request, pk):
        # Get original student
        aluno_original = get_object_or_404(Aluno, pk=pk)
        
        user = request.user
        if not hasattr(user, 'profile') or not user.profile.escola:
             messages.error(request, "Você não está vinculado a nenhuma escola para importar o aluno.")
             return redirect('alunos:verificar_cpf')

        escola_destino = user.profile.escola
        
        # Double check if already exists (prevent race condition)
        if Aluno.objects.filter(escola=escola_destino, cpf=aluno_original.cpf).exists():
            messages.warning(request, "Este aluno já foi importado ou cadastrado nesta escola.")
            # Find the one that exists to redirect to it? Or just list?
            # Let's redirect to list for safety
            return redirect('alunos:lista_alunos')

        # Clone data
        novo_aluno = Aluno(
            escola=escola_destino,
            nome_completo=aluno_original.nome_completo,
            cpf=aluno_original.cpf,
            rg=aluno_original.rg,
            orgao_exp=aluno_original.orgao_exp,
            data_emissao=aluno_original.data_emissao,
            data_nascimento=aluno_original.data_nascimento,
            sexo=aluno_original.sexo,
            estado_civil=aluno_original.estado_civil,
            cor_raca=aluno_original.cor_raca,
            nome_mae=aluno_original.nome_mae,
            naturalidade=aluno_original.naturalidade,
            uf_naturalidade=aluno_original.uf_naturalidade,
            deficiencia=aluno_original.deficiencia,
            escolaridade=aluno_original.escolaridade,
            email_principal=aluno_original.email_principal, # Note: Emails are unique=True in model? 
            # Wait, email_principal is unique=True in model. 
            # If we clone, we will have duplicate email error!
            # Logic adjustment needed below.
            whatsapp=aluno_original.whatsapp,
            telefone_principal=aluno_original.telefone_principal,
            endereco_cep=aluno_original.endereco_cep,
            endereco_rua=aluno_original.endereco_rua,
            endereco_numero=aluno_original.endereco_numero,
            endereco_bairro=aluno_original.endereco_bairro,
            endereco_cidade=aluno_original.endereco_cidade,
            endereco_estado=aluno_original.endereco_estado,
            tempo_moradia=aluno_original.tempo_moradia,
            tipo_moradia=aluno_original.tipo_moradia,
            valor_moradia=aluno_original.valor_moradia,
            situacao_profissional=aluno_original.situacao_profissional,
            renda_individual=aluno_original.renda_individual,
            num_moradores=aluno_original.num_moradores,
            quantos_trabalham=aluno_original.quantos_trabalham,
            renda_moradores=aluno_original.renda_moradores,
            como_soube=aluno_original.como_soube
        )
        
        # Handle Email Uniqueness for Cloning
        # Email is no longer unique globally, so we can copy it directly.
        
        try:
            novo_aluno.save()
            
            # Clonar os Arquivos Digitais (Pasta Digital)
            from .models import ArquivoAluno
            arquivos_originais = aluno_original.arquivos.all()
            for arq_orig in arquivos_originais:
                # Criar novo arquivo copiando o conteúdo
                novo_arquivo = ArquivoAluno(
                    aluno=novo_aluno,
                    nome=arq_orig.nome,
                    enviado_por=request.user
                )
                # Copiar o arquivo físico
                if arq_orig.arquivo:
                    from django.core.files.base import ContentFile
                    novo_arquivo.arquivo.save(
                        os.path.basename(arq_orig.arquivo.name),
                        ContentFile(arq_orig.arquivo.read()),
                        save=False
                    )
                novo_arquivo.save()

            # Log manual da clonagem para acionar WebSocket
            self.save_log(novo_aluno, 'CREATE', {'origem_clonagem': str(aluno_original.pk), 'acao': 'clonagem'})
            
            messages.success(request, f"Aluno {novo_aluno.nome_completo} importado com sucesso! Os documentos digitais também foram transferidos.")
            return redirect('alunos:editar_aluno', pk=novo_aluno.pk)
        except Exception as e:
            messages.error(request, f"Erro ao importar aluno: {e}")
            return redirect('alunos:verificar_cpf')


class AlunoListView(LoginRequiredMixin, ListView):
    model = Aluno
    template_name = 'alunos/aluno_list.html'
    context_object_name = 'alunos'
    paginate_by = 20 # Default pagination

    def get_queryset(self):
        user = self.request.user
        queryset = Aluno.objects.all().order_by('-data_criacao', 'nome_completo') # Order by enrollment date/time

        if user.is_superuser:
            escola_filter = self.request.GET.get('escola')
            if escola_filter:
                queryset = queryset.filter(escola__id=escola_filter)
        elif hasattr(user, 'profile') and user.profile.escola:
            queryset = queryset.filter(escola=user.profile.escola)
        else:
            return Aluno.objects.none()

        # Search functionality
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(nome_completo__icontains=search_query) | 
                Q(cpf__icontains=search_query)
            )
        
        # Filtro por Tipo de Curso (Cursos de Interesse)
        tipo_curso_id = self.request.GET.get('tipo_curso')
        if tipo_curso_id:
            queryset = queryset.filter(cursos_interesse__id=tipo_curso_id)
        
        # Anotar o total de arquivos que o aluno tem
        queryset = queryset.annotate(
            total_arquivos=Count('arquivos')
        )
        
        # Implement dynamic pagination
        page_size = self.request.GET.get('page_size', self.paginate_by)
        try:
            self.paginate_by = int(page_size)
        except ValueError:
            self.paginate_by = 20 # Fallback to default if invalid
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        if user.is_superuser:
            context['todas_escolas'] = Escola.objects.all().order_by('nome')
            context['escola_selecionada'] = self.request.GET.get('escola', '')
            context['todos_cursos_interesse'] = TipoCurso.objects.all().order_by('nome')
        elif hasattr(user, 'profile') and user.profile.escola:
            context['todos_cursos_interesse'] = TipoCurso.objects.filter(escola=user.profile.escola).order_by('nome')
        
        context['tipo_curso_selecionado'] = self.request.GET.get('tipo_curso', '')
        
        return context

class AlunoDetailView(StaffRequiredMixin, DetailView):
    model = Aluno
    template_name = 'alunos/aluno_detail.html'
    context_object_name = 'aluno'

class AlunoCreateView(AuditLogMixin, StaffRequiredMixin, CreateView):
    model = Aluno
    form_class = AlunoForm
    template_name = 'alunos/aluno_form.html'
    
    def get_success_url(self):
        return reverse_lazy('alunos:cadastro_sucesso', kwargs={'pk': self.object.pk})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_initial(self):
        initial = super().get_initial()
        cpf = self.request.GET.get('cpf')
        if cpf:
            initial['cpf'] = cpf
        return initial

    def form_valid(self, form):
        if not self.request.user.is_superuser:
            form.instance.escola = self.request.user.profile.escola
        return super().form_valid(form)

class AlunoCadastroSucessoView(StaffRequiredMixin, DetailView):
    model = Aluno
    template_name = 'alunos/cadastro_sucesso.html'
    context_object_name = 'aluno'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        aluno = self.object
        
        # Mensagem para o WhatsApp
        nome_escola = aluno.escola.nome if aluno.escola else "Escola"
        mensagem = f"Sua Inscrição foi realizada com sucesso! Aguarde o nosso contato para os cursos em que você se inscreveu!! Att: {nome_escola}"
        mensagem_encoded = quote(mensagem)
        
        # Gera o link se o aluno tiver número válido
        if aluno.whatsapp_link:
             context['whatsapp_link'] = f"https://wa.me/{aluno.whatsapp_link}?text={mensagem_encoded}"
        
        return context

class AlunoUpdateView(AuditLogMixin, StaffRequiredMixin, UpdateView):
    model = Aluno
    template_name = 'alunos/aluno_form.html'
    success_url = reverse_lazy('alunos:lista_alunos')

    def get_form_class(self):
        """
        Retorna o formulário apropriado com base no grupo do usuário.
        """
        if self.request.user.groups.filter(name='Auxiliar Administrativo').exists():
            return AuxiliarAlunoForm
        return AlunoForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

class AlunoDeleteView(AuditLogMixin, StaffRequiredMixin, DeleteView):

    model = Aluno

    template_name = 'alunos/aluno_confirm_delete.html'

    success_url = reverse_lazy('alunos:lista_alunos')



    def test_func(self):

        # Sobrescreve o test_func do StaffRequiredMixin para impedir que Auxiliares excluam

        if self.request.user.groups.filter(name='Auxiliar Administrativo').exists():

            return False

        return super().test_func()



class AlunoHistoricoView(StaffRequiredMixin, DetailView):
    model = Aluno
    template_name = 'alunos/aluno_historico.html'
    context_object_name = 'aluno'

    def get_template_names(self):
        if self.request.GET.get('popup'):
            return ['alunos/aluno_historico_popup.html']
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        aluno = self.object
        
        # Inscrições atuais com avaliações e chamadas (Dossiê)
        from django.db.models import Prefetch
        from cursos.models import Chamada
        
        inscricoes = aluno.inscricao_set.all().select_related('curso__escola').prefetch_related(
            'avaliacao_professor',
            'avaliacao_aluno',
            Prefetch('chamadas', queryset=Chamada.objects.select_related('registro_aula').order_by('registro_aula__data_aula'))
        )
        context['inscricoes'] = inscricoes

        # Logs de mudança de status (Historico de Deferimentos/Desistências)
        from django.contrib.contenttypes.models import ContentType
        from core.models import AuditLog
        from cursos.models import Inscricao
        
        insc_ct = ContentType.objects.get_for_model(Inscricao)
        insc_ids = [str(i.id) for i in inscricoes]
        
        logs_status_raw = AuditLog.objects.filter(
            content_type=insc_ct,
            object_id__in=insc_ids,
            acao='UPDATE'
        ).select_related('usuario', 'content_type').order_by('-data_hora')

        # Extraír apenas mudanças de status
        timeline = []
        for log in logs_status_raw:
            try:
                details = json.loads(log.detalhes)
                if 'alteracoes' in details and 'status' in details['alteracoes']:
                    novo_status = details['alteracoes']['status']
                    # Mapear status para texto legível
                    from cursos.models import Inscricao
                    status_text = dict(Inscricao.STATUS_CHOICES).get(novo_status, novo_status)
                    
                    timeline.append({
                        'data_hora': log.data_hora,
                        'usuario': log.usuario,
                        'status': status_text,
                        'inscricao_id': log.object_id
                    })
            except:
                continue

        context['timeline'] = timeline
        return context

class AlunoUpdateObservacoesView(AuditLogMixin, StaffRequiredMixin, View):
    def post(self, request, pk):
        aluno = get_object_or_404(Aluno, pk=pk)
        observacoes = request.POST.get('observacoes')
        from core.utils import audit_context
        with audit_context(skip=True):
            aluno.observacoes = observacoes
            aluno.save()
        
        # Log manual
        self.save_log(aluno, 'UPDATE', {'campo': 'observacoes', 'valor': 'atualizado'})
        
        messages.success(request, "Observações do histórico atualizadas com sucesso.")
        return redirect('alunos:historico_aluno', pk=aluno.pk)

class AlunoUpdateCursosInteresseView(AuditLogMixin, StaffRequiredMixin, View):
    def post(self, request, pk):
        aluno = get_object_or_404(Aluno, pk=pk)
        cursos_ids = request.POST.getlist('cursos_interesse')
        
        # Validar se os cursos pertencem à mesma escola ou se é superuser
        user = request.user
        if not user.is_superuser:
            if hasattr(user, 'profile') and user.profile.escola:
                # Filtrar cursos_ids para garantir que só salve os da escola dele
                valid_ids = list(TipoCurso.objects.filter(
                    id__in=cursos_ids, 
                    escola=user.profile.escola
                ).values_list('id', flat=True))
                cursos_ids = valid_ids
        
        from core.utils import audit_context
        with audit_context(skip=True):
            aluno.cursos_interesse.set(cursos_ids)
            aluno.save() # Força atualização do timestamp 'data_atualizacao'
        
        # Log manual
        self.save_log(aluno, 'UPDATE', {'campo': 'cursos_interesse', 'total': len(cursos_ids)})
        
        messages.success(request, f"Cursos de interesse de {aluno.nome_completo} atualizados.")
        return redirect('alunos:lista_alunos')


from core.models import AuditLog # Import AuditLog

class AlunoCSVUploadView(LoginRequiredMixin, SuperuserRequiredMixin, View): # Apenas superusuário pode fazer upload
    template_name = 'alunos/upload_alunos_csv.html'

    def get_context_data(self, form, sample_csv_headers, sample_csv_content):
        """Helper method to build common context for GET and POST requests"""
        context = {
            'form': form,
            'sample_csv_headers': sample_csv_headers,
            'sample_csv_content': sample_csv_content,
            'Aluno': Aluno, # Passa o modelo Aluno para acessar choices no template
        }
        return context

    def handle_uploaded_file(self, file):
        decoded_file = file.read().decode('cp1252')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        
        created_count = 0
        updated_count = 0
        errors = []

        # Obtém os valores de choices para validação
        SEXO_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.SEXO_CHOICES}
        ESTADO_CIVIL_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.ESTADO_CIVIL_CHOICES}
        COR_RACA_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.COR_RACA_CHOICES}
        ESCOLARIDADE_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.ESCOLARIDADE_CHOICES}
        SITUACAO_PROFISSIONAL_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.SITUACAO_PROFISSIONAL_CHOICES}
        TEMPO_MORADIA_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.TEMPO_MORADIA_CHOICES}
        TIPO_MORADIA_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.TIPO_MORADIA_CHOICES}
        COMO_SOUBE_CHOICES = {choice[1].lower(): choice[0] for choice in Aluno.COMO_SOUBE_CHOICES}


        for i, row in enumerate(reader):
            line_num = i + 2 # +1 for 0-indexed, +1 for header
            try:
                # Campo 'escola_nome' é obrigatório no CSV para vincular o aluno a uma escola existente
                escola_nome = row.get('escola_nome')
                if not escola_nome:
                    raise ValueError("Coluna 'escola_nome' é obrigatória.")
                try:
                    escola = Escola.objects.get(nome=escola_nome)
                except Escola.DoesNotExist:
                    raise ValueError(f"Escola '{escola_nome}' não encontrada. A escola deve existir previamente.")

                # Mapeamento e Conversão dos dados
                aluno_data = {
                    'escola': escola,
                    'nome_completo': row.get('Nome'),
                    'cpf': row.get('cpf'),
                    'rg': row.get('rg'),
                    'orgao_exp': row.get('orgaoemissor'),
                    'naturalidade': row.get('Naturalidade'),
                    'uf_naturalidade': row.get('uf'),
                    'endereco_cep': row.get('cep'),
                    'endereco_rua': row.get('endereco'),
                    'endereco_numero': row.get('num'),
                    'endereco_bairro': row.get('bairro'),
                    'endereco_cidade': row.get('cidade'), # Novo
                    'endereco_estado': row.get('estado_endereco'), # Novo
                    'telefone_principal': row.get('telefone1'),
                    'email_principal': row.get('email'), # Novo
                    'whatsapp': '', # Default para vazio

                    'valor_moradia': float(row.get('valorresidencia')) if row.get('valorresidencia') else 0.0,
                    'renda_individual': float(row.get('renda')) if row.get('renda') else 0.0,
                    'num_moradores': int(row.get('nummoradores')) if row.get('nummoradores') else 0,
                    'quantos_trabalham': int(row.get('numtrabalham')) if row.get('numtrabalham') else 0,
                    'renda_moradores': float(row.get('rendatotal')) if row.get('rendatotal') else 0.0,
                }

                # Conversão de datas (DD/MM/YYYY)
                data_nasc_str = row.get('Data nasc (00/00/0000)')
                if data_nasc_str:
                    aluno_data['data_nascimento'] = datetime.strptime(data_nasc_str, '%d/%m/%Y').date()
                data_emissao_str = row.get('dataemissao')
                if data_emissao_str:
                    aluno_data['data_emissao'] = datetime.strptime(data_emissao_str, '%d/%m/%Y').date()
                
                # Conversão de booleano (deficiencia) a partir de uma string Sim/Nao
                deficiencia_str = row.get('deficiencia (sim/nao)', 'nao').lower() # Padrão para 'nao'
                aluno_data['deficiencia'] = (deficiencia_str == 'sim')

                # Mapeamento de CHOICES (case-insensitive para os valores em português)
                aluno_data['cor_raca'] = COR_RACA_CHOICES.get(row.get('Cor', '').lower())
                aluno_data['sexo'] = SEXO_CHOICES.get(row.get('Sexo', '').lower())
                aluno_data['estado_civil'] = ESTADO_CIVIL_CHOICES.get(row.get('Estado Civil', '').lower())
                aluno_data['escolaridade'] = ESCOLARIDADE_CHOICES.get(row.get('escolaridade', '').lower())
                aluno_data['situacao_profissional'] = SITUACAO_PROFISSIONAL_CHOICES.get(row.get('situacao', '').lower())

                # Campos novos com choices
                aluno_data['tempo_moradia'] = TEMPO_MORADIA_CHOICES.get(row.get('tempo_moradia', '').lower()) 
                aluno_data['tipo_moradia'] = TIPO_MORADIA_CHOICES.get(row.get('tipo_moradia', '').lower())
                aluno_data['como_soube'] = COMO_SOUBE_CHOICES.get(row.get('como_soube', '').lower())
                
                # Remoção de campos vazios ou None para que os defaults do modelo sejam aplicados
                aluno_data = {k: v for k, v in aluno_data.items() if v is not None and v != ''}


                # Validação de campos obrigatórios no modelo Aluno que não têm default
                required_fields_to_check = ['nome_completo', 'cpf', 'data_nascimento', 'sexo', 'escola', 'email_principal', 'endereco_cidade', 'endereco_estado']
                for field_name in required_fields_to_check:
                    if field_name not in aluno_data or not aluno_data[field_name]:
                        # Ajusta a mensagem de erro para ser mais específica
                        if field_name == 'cor_raca' and row.get('Cor', '') and not aluno_data['cor_raca']:
                            raise ValueError(f"Valor inválido para 'Cor': {row.get('Cor')}. Opções válidas: {list(COR_RACA_CHOICES.keys())}")
                        elif field_name == 'sexo' and row.get('Sexo', '') and not aluno_data['sexo']:
                            raise ValueError(f"Valor inválido para 'Sexo': {row.get('Sexo')}. Opções válidas: {list(SEXO_CHOICES.keys())}")
                        elif field_name == 'estado_civil' and row.get('Estado Civil', '') and not aluno_data['estado_civil']:
                            raise ValueError(f"Valor inválido para 'Estado Civil': {row.get('Estado Civil')}. Opções válidas: {list(ESTADO_CIVIL_CHOICES.keys())}")
                        elif field_name == 'escolaridade' and row.get('escolaridade', '') and not aluno_data['escolaridade']:
                            raise ValueError(f"Valor inválido para 'escolaridade': {row.get('escolaridade')}. Opções válidas: {list(ESCOLARIDADE_CHOICES.keys())}")
                        elif field_name == 'situacao_profissional' and row.get('situacao', '') and not aluno_data['situacao_profissional']:
                            raise ValueError(f"Valor inválido para 'situacao': {row.get('situacao')}. Opções válidas: {list(SITUACAO_PROFISSIONAL_CHOICES.keys())}")
                        else:
                            raise ValueError(f"Campo obrigatório '{field_name}' está faltando ou é inválido.")


                # Tenta encontrar um aluno existente pelo CPF (assumindo CPF como único identificador)
                aluno_instance = Aluno.objects.filter(cpf=aluno_data['cpf']).first()

                if aluno_instance:
                    # Se o aluno existe, atualiza os dados
                    for key, value in aluno_data.items():
                        setattr(aluno_instance, key, value)
                    aluno_instance.save()
                    updated_count += 1
                else:
                    # Se o aluno não existe, cria um novo
                    Aluno.objects.create(**aluno_data)
                    created_count += 1

            except (ValueError, KeyError, ValidationError, Escola.DoesNotExist) as e:
                errors.append(f"Linha {line_num}: {e}")
            except Exception as e:
                errors.append(f"Linha {line_num}: Erro inesperado - {e}")
        
        return created_count, updated_count, errors

    def get(self, request, *args, **kwargs):
        form = AlunoCSVUploadForm()
        # Aqui, podemos passar um exemplo de CSV para o template
        # Certifique-se de que a escola_nome exista no seu banco de dados para o exemplo
        sample_csv_headers = "escola_nome,Nome,Data nasc (00/00/0000),Cor,Sexo,Estado Civil,Filiacao,Naturalidade,uf,cep,endereco,num,bairro,cidade,estado_endereco,telefone1,email,rg,orgaoemissor,dataemissao,cpf,deficiencia (sim/nao),escolaridade,situacao,renda,nummoradores,numtrabalham,rendatotal,valorresidencia,tempo_moradia,tipo_moradia,como_soube"
        sample_csv_data = [
            "CP Luizote,Maria da Silva,15/03/1990,Parda,Feminino,Casado,Ana da Silva,Uberlândia,MG,38400-000,Rua das Flores,123,Centro,Uberlândia,MG,(34)99999-9999,maria@email.com,1234567,SSP,01/01/2010,123.456.789-01,nao,Ensino Médio Completo,Estudante,1500.00,3,1,800.00,500.00,Mais de 5 anos,Alugada,Redes Sociais"
        ]
        sample_csv_content = sample_csv_headers + "\n" + "\n".join(sample_csv_data)
        
        context = self.get_context_data(form, sample_csv_headers, sample_csv_content)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = AlunoCSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            if not csv_file.name.endswith('.csv'):
                messages.error(request, "Por favor, envie um arquivo CSV válido.")
                # Renderiza o formulário novamente com o contexto do GET para exibir o exemplo CSV
                sample_csv_headers = "escola_nome,Nome,Data nasc (00/00/0000),Cor,Sexo,Estado Civil,Filiacao,Naturalidade,uf,cep,endereco,num,bairro,cidade,estado_endereco,telefone1,email,rg,orgaoemissor,dataemissao,cpf,deficiencia (sim/nao),escolaridade,situacao,renda,nummoradores,numtrabalham,rendatotal,valorresidencia,tempo_moradia,tipo_moradia,como_soube"
                sample_csv_data = [
                    "CP Luizote,Maria da Silva,15/03/1990,Parda,Feminino,Casado,Ana da Silva,Uberlândia,MG,38400-000,Rua das Flores,123,Centro,Uberlândia,MG,(34)99999-9999,maria@email.com,1234567,SSP,01/01/2010,123.456.789-01,nao,Ensino Médio Completo,Estudante,1500.00,3,1,800.00,500.00,Mais de 5 anos,Alugada,Redes Sociais"
                ]
                sample_csv_content = sample_csv_headers + "\n" + "\n".join(sample_csv_data)
                context = self.get_context_data(form, sample_csv_headers, sample_csv_content)
                return render(request, self.template_name, context)

            created_count, updated_count, errors = self.handle_uploaded_file(csv_file)

            if errors:
                for error in errors:
                    messages.error(request, error)
                messages.warning(request, f"Processamento concluído com {created_count} alunos criados, {updated_count} alunos atualizados e erros em algumas linhas.")
            else:
                messages.success(request, f"Upload de CSV concluído com sucesso: {created_count} alunos criados, {updated_count} alunos atualizados.")
            
            # Log da ação em massa para WebSocket
            try:
                AuditLog.objects.create(
                    usuario=request.user,
                    acao='CREATE',
                    detalhes=f"Upload CSV Alunos: {created_count} criados, {updated_count} atualizados.",
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except Exception as e:
                print(f"Erro log CSV Aluno: {e}")

            return redirect(reverse_lazy('alunos:lista_alunos'))
        else:
            # Se o formulário não for válido (ex: nenhum arquivo enviado), renderiza com erros
            sample_csv_headers = "escola_nome,Nome,Data nasc (00/00/0000),Cor,Sexo,Estado Civil,Filiacao,Naturalidade,uf,cep,endereco,num,bairro,cidade,estado_endereco,telefone1,email,rg,orgaoemissor,dataemissao,cpf,deficiencia (sim/nao),escolaridade,situacao,renda,nummoradores,numtrabalham,rendatotal,valorresidencia,tempo_moradia,tipo_moradia,como_soube"
            sample_csv_data = [
                "CP Luizote,Maria da Silva,15/03/1990,Parda,Feminino,Casado,Ana da Silva,Uberlândia,MG,38400-000,Rua das Flores,123,Centro,Uberlândia,MG,(34)99999-9999,maria@email.com,1234567,SSP,01/01/2010,123.456.789-01,nao,Ensino Médio Completo,Estudante,1500.00,3,1,800.00,500.00,Mais de 5 anos,Alugada,Redes Sociais"
            ]
            sample_csv_content = sample_csv_headers + "\n" + "\n".join(sample_csv_data)
            context = self.get_context_data(form, sample_csv_headers, sample_csv_content)
            return render(request, self.template_name, context)


class AlunoArquivoAjaxUploadView(LoginRequiredMixin, StaffRequiredMixin, View):
    def post(self, request, pk):
        aluno = get_object_or_404(Aluno, pk=pk)
        arquivo = request.FILES.get('arquivo')
        nome = request.POST.get('nome', '')

        if not arquivo:
            return JsonResponse({'sucesso': False, 'erro': 'Nenhum arquivo enviado.'}, status=400)

        from .models import ArquivoAluno
        doc = ArquivoAluno.objects.create(
            aluno=aluno,
            arquivo=arquivo,
            nome=nome or arquivo.name,
            enviado_por=request.user
        )

        return JsonResponse({
            'sucesso': True,
            'arquivo': {
                'id': doc.id,
                'nome': doc.nome,
                'url': doc.arquivo.url,
                'data': doc.data_upload.strftime('%d/%m/%Y'),
                'is_pdf': doc.is_pdf,
                'is_image': doc.is_image
            }
        })


class AlunoArquivoActionView(LoginRequiredMixin, StaffRequiredMixin, View):
    def delete(self, request, pk, file_id):
        from .models import ArquivoAluno
        arquivo_obj = get_object_or_404(ArquivoAluno, id=file_id, aluno_id=pk)
        
        # Remover arquivo físico
        if arquivo_obj.arquivo:
            if os.path.isfile(arquivo_obj.arquivo.path):
                os.remove(arquivo_obj.arquivo.path)
        
        arquivo_obj.delete()
        return JsonResponse({'sucesso': True})

    def post(self, request, pk, file_id):
        from .models import ArquivoAluno
        arquivo_obj = get_object_or_404(ArquivoAluno, id=file_id, aluno_id=pk)
        novo_nome = request.POST.get('novo_nome')
        if novo_nome:
            arquivo_obj.nome = novo_nome
            arquivo_obj.save()
            return JsonResponse({'sucesso': True, 'nome': arquivo_obj.nome})
        return JsonResponse({'sucesso': False, 'erro': 'Nome inválido.'}, status=400)

    def get(self, request, pk):
        aluno = get_object_or_404(Aluno, pk=pk)
        arquivos = aluno.arquivos.all()
        data = []
        for a in arquivos:
            data.append({
                'id': a.id,
                'nome': a.nome,
                'url': a.arquivo.url,
                'data': a.data_upload.strftime('%d/%m/%Y'),
                'is_pdf': a.is_pdf,
                'is_image': a.is_image
            })
        return JsonResponse({'sucesso': True, 'arquivos': data})


class WebSocialListView(LoginRequiredMixin, SuperuserRequiredMixin, ListView):
    model = WebSocialMember
    template_name = 'alunos/web_social_list.html'
    context_object_name = 'membros'
    paginate_by = 50

    def get_queryset(self):
        queryset = WebSocialMember.objects.select_related('aluno', 'aluno__escola').all()
        
        # Filtro de Ano
        self.ano_atual = timezone.now().year
        self.ano_selecionado = self.request.GET.get('ano', str(self.ano_atual))
        
        try:
            queryset = queryset.filter(ano_inclusao=int(self.ano_selecionado))
        except (ValueError, TypeError):
            queryset = queryset.filter(ano_inclusao=self.ano_atual)
            self.ano_selecionado = str(self.ano_atual)
            
        # Busca por Nome ou CPF
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(aluno__nome_completo__icontains=search_query) | 
                Q(aluno__cpf__icontains=search_query)
            )
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ano_selecionado'] = int(self.ano_selecionado)
        
        # Lista de anos disponíveis para o filtro
        anos = WebSocialMember.objects.values_list('ano_inclusao', flat=True).distinct().order_by('-ano_inclusao')
        anos_list = list(anos)
        if self.ano_atual not in anos_list:
            anos_list.insert(0, self.ano_atual)
        context['anos_disponiveis'] = sorted(list(set(anos_list)), reverse=True)
        
        return context


class WebSocialExportExcelView(LoginRequiredMixin, SuperuserRequiredMixin, View):
    def get(self, request):
        # Aplicar os mesmos filtros da AlunoListView
        queryset = WebSocialMember.objects.select_related('aluno', 'aluno__escola').all()
        
        ano_selecionado = request.GET.get('ano')
        if ano_selecionado:
            try:
                queryset = queryset.filter(ano_inclusao=int(ano_selecionado))
            except (ValueError, TypeError):
                pass
                
        search_query = request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(aluno__nome_completo__icontains=search_query) | 
                Q(aluno__cpf__icontains=search_query)
            )

        # Criar o workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Web Social {ano_selecionado or ''}"

        # Cabeçalhos
        columns = [
            'Ano Inclusão', 'Data Inclusão', 'Escola', 'Nome Completo', 'CPF', 'RG', 'Órgão Exp', 
            'Data Emissão', 'Data Nascimento', 'Sexo', 'Estado Civil', 'Cor/Raça', 
            'Nome da Mãe', 'Naturalidade', 'UF Naturalidade', 'Escolaridade', 
            'Possui Deficiência', 'Tipo de Deficiência'
        ]
        ws.append(columns)

        # Estilizar cabeçalho
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Adicionar dados
        for membro in queryset:
            aluno = membro.aluno
            ws.append([
                membro.ano_inclusao,
                membro.data_inclusao.strftime('%d/%m/%Y %H:%M') if membro.data_inclusao else '',
                aluno.escola.nome if aluno.escola else '',
                aluno.nome_completo,
                aluno.cpf,
                aluno.rg or '',
                aluno.orgao_exp or '',
                aluno.data_emissao.strftime('%d/%m/%Y') if aluno.data_emissao else '',
                aluno.data_nascimento.strftime('%d/%m/%Y') if aluno.data_nascimento else '',
                aluno.sexo,
                aluno.estado_civil,
                aluno.cor_raca or '',
                aluno.nome_mae or '',
                aluno.naturalidade or '',
                aluno.uf_naturalidade or '',
                aluno.escolaridade or '',
                'Sim' if aluno.deficiencia else 'Não',
                aluno.tipo_deficiencia or ''
            ])

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Preparar resposta
        filename = f"web_social_{ano_selecionado or 'geral'}_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
